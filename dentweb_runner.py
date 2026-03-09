"""이미지 인식 기반 DentWeb 자동화 시퀀스

전체 흐름:
1. 덴트웹 창 자동 탐색 → 활성화 (최소화 복원 포함)
2. 경영/통계 → 임플란트 수술 통계
3. 특정기간 → 부터 '오늘' → 까지 '오늘' (당일 조회)
4. 엑셀저장 → 저장 다이얼로그 → 파일 저장
5. "수술기록지" 시트 2행 이하 데이터 유무로 판별

학습 모드에서 각 버튼의 스크린샷을 캡처하고,
자동화 시 pyautogui.locateCenterOnScreen()으로 이미지를 찾아 클릭.
"""

import os
import json
import time
import glob
import pyautogui
import pyperclip
from PIL import Image
from openpyxl import load_workbook


STEPS_FILE = "dentweb_steps.json"
TEMPLATES_DIR = "templates"
WINDOW_TITLE = "덴트웹"

# 템플릿 캡처 크기 (마우스 위치 중심으로 캡처)
CAPTURE_W = 120
CAPTURE_H = 50


def _paste_text(text: str):
    """클립보드에 복사 후 Ctrl+V로 붙여넣기 (한글 지원)"""
    pyperclip.copy(text)
    time.sleep(0.05)
    pyautogui.hotkey("ctrl", "v")
    time.sleep(0.1)


def _win32_click(x: int, y: int, activate_first: bool = False):
    """왼쪽 클릭. activate_first=True면 활성화용 클릭 1회 추가"""
    ix, iy = int(x), int(y)
    if activate_first:
        pyautogui.click(ix, iy, button="left")  # 창 활성화용
        time.sleep(0.3)
    pyautogui.click(ix, iy, button="left")  # 실제 클릭
    time.sleep(0.1)


# --- 클릭 시퀀스 정의 ---

DATA_STEPS = [
    {"name": "stats_menu", "label": "상단 '경영/통계' 아이콘", "x": None, "y": None, "wait_after": 3.0},
    {"name": "implant_tab", "label": "왼쪽 사이드바 '임플란트 수술 통계'", "x": None, "y": None, "wait_after": 3.0},
    {"name": "custom_period", "label": "'특정기간' 라디오 버튼", "x": None, "y": None, "wait_after": 1.5},
    {"name": "date_start_field", "label": "'부터' 날짜 필드 클릭 (달력 열기)", "x": None, "y": None, "wait_after": 1.5},
    {"name": "date_start_today", "label": "'부터' 달력 하단 '오늘' 버튼", "x": None, "y": None, "wait_after": 1.5},
    {"name": "date_end_field", "label": "'까지' 날짜 필드 클릭 (달력 열기)", "x": None, "y": None, "wait_after": 1.5},
    {"name": "date_end_today", "label": "'까지' 달력 하단 '오늘' 버튼 (자동 조회됨)", "x": None, "y": None, "wait_after": 5.0},
    {"name": "export_btn", "label": "'엑셀저장' 버튼", "x": None, "y": None, "wait_after": 3.0},
    # ── 저장 다이얼로그 단계 ──────────────────────────────────────
    # 안내: 엑셀저장 버튼 클릭 후 '다른 이름으로 저장' 창이 열린 상태에서 아래 좌표를 설정하세요.
    {"name": "save_dialog_agent_folder",  "label": "저장 창 — '덴트웹 에이전트' 폴더 위치 (더블클릭 진입)", "x": None, "y": None, "wait_after": 1.5},
    {"name": "save_dialog_exports_folder","label": "저장 창 — 'exports' 폴더 위치 (더블클릭 진입)", "x": None, "y": None, "wait_after": 1.5},
    {"name": "save_dialog_filename_field","label": "저장 창 — 파일 이름 입력 필드 (클릭 후 파일명 입력됨)", "x": None, "y": None, "wait_after": 0.5},
    {"name": "save_dialog_save_btn",      "label": "저장 창 — '저장(S)' 버튼", "x": None, "y": None, "wait_after": 2.0},
    {"name": "save_dialog_confirm_yes",   "label": "덮어쓰기 확인 팝업 — '예(Y)' 버튼", "x": None, "y": None, "wait_after": 2.0},
]


def _get_template_path(step_name: str) -> str:
    return os.path.join(TEMPLATES_DIR, f"{step_name}.png")


def _capture_template(x: int, y: int, step_name: str) -> str:
    """마우스 위치 주변 영역을 스크린샷으로 캡처하여 템플릿 저장"""
    os.makedirs(TEMPLATES_DIR, exist_ok=True)

    # 캡처 영역 계산 (화면 경계 처리)
    screen_w, screen_h = pyautogui.size()
    left = max(0, x - CAPTURE_W // 2)
    top = max(0, y - CAPTURE_H // 2)
    right = min(screen_w, left + CAPTURE_W)
    bottom = min(screen_h, top + CAPTURE_H)
    w = right - left
    h = bottom - top

    screenshot = pyautogui.screenshot(region=(left, top, w, h))
    path = _get_template_path(step_name)
    screenshot.save(path)
    return path


def _find_and_click(step: dict, log_callback=None, confidence: float = 0.8,
                    activate_first: bool = False) -> bool:
    """이미지 인식으로 버튼을 찾아 클릭. 실패 시 좌표 폴백."""
    def _log(msg):
        if log_callback:
            log_callback(msg)

    template_path = _get_template_path(step["name"])

    # 1차: 이미지 인식
    if os.path.exists(template_path):
        try:
            location = pyautogui.locateCenterOnScreen(
                template_path, confidence=confidence
            )
            if location:
                cx, cy = location
                _log(f"이미지 발견: {step['label']} → 클릭 ({cx}, {cy})")
                _win32_click(cx, cy, activate_first=activate_first)
                return True
            else:
                _log(f"이미지 못 찾음: {step['label']} — 좌표 폴백 시도")
        except Exception as e:
            _log(f"이미지 검색 오류: {e} — 좌표 폴백 시도")

    # 2차: 저장된 좌표로 폴백
    if step.get("x") is not None and step.get("y") is not None:
        _log(f"좌표 폴백: {step['label']} ({step['x']}, {step['y']})")
        _win32_click(step["x"], step["y"], activate_first=activate_first)
        return True

    _log(f"클릭 실패: {step['label']} — 이미지도 좌표도 없음")
    return False


# --- 설정 파일 관리 ---

def load_config_data(path: str = STEPS_FILE) -> dict | None:
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, list):
        return None

    # 저장 다이얼로그 단계는 선택 사항 — 미설정이어도 run 가능 (폴백 처리)
    OPTIONAL_STEPS = {
        "save_dialog_agent_folder", "save_dialog_exports_folder",
        "save_dialog_filename_field", "save_dialog_save_btn", "save_dialog_confirm_yes",
    }

    for step in data.get("data_steps", []):
        if step.get("skip"):
            continue
        if step.get("name") == "data_check":
            continue
        if step.get("name") in OPTIONAL_STEPS:
            continue
        # 이미지 템플릿이 있으면 좌표 없어도 OK
        template_path = _get_template_path(step.get("name", ""))
        if os.path.exists(template_path):
            continue
        if step.get("x") is None or step.get("y") is None:
            return None
    return data


def save_config_data(data: dict, path: str = STEPS_FILE):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# --- 학습 모드 (CLI - 레거시) ---

CAPTURE_DELAY = 5


def _countdown_capture() -> tuple[int, int]:
    for remaining in range(CAPTURE_DELAY, 0, -1):
        print(f"\r  → {remaining}초 후 마우스 위치 저장...", end="", flush=True)
        time.sleep(1)
    x, y = pyautogui.position()
    print(f"\r  → 저장됨: ({x}, {y})              ")
    return x, y


def run_teach_mode() -> dict:
    print()
    print("=" * 55)
    print("  덴트웹 자동화 - 클릭 위치 학습")
    print("=" * 55)
    input("준비되면 Enter...")

    steps = json.loads(json.dumps(DATA_STEPS))
    i = 0
    while i < len(steps):
        step = steps[i]
        print(f"\n[{i + 1}/{len(steps)}] {step['label']}")
        choice = input("  → Enter/s/r: ").strip().lower()
        if choice == "r":
            steps = json.loads(json.dumps(DATA_STEPS))
            i = 0
            continue
        if choice == "s":
            steps[i]["skip"] = True
            i += 1
            continue
        x, y = _countdown_capture()
        steps[i]["x"] = x
        steps[i]["y"] = y
        steps[i]["skip"] = False
        _capture_template(x, y, step["name"])
        i += 1

    result = {"data_steps": steps}
    save_config_data(result)
    return result


# --- 메인 Runner ---

class DentwebRunner:
    def __init__(self, cfg: dict):
        self.download_dir = cfg["download_dir"]
        self.download_timeout = cfg.get("download_timeout_seconds", 30)
        self._data = load_config_data()

    def is_configured(self) -> bool:
        return self._data is not None

    def _get_save_step(self, name: str) -> dict | None:
        """저장 다이얼로그용 단계 가져오기 (좌표 설정된 경우만)"""
        for step in self._data.get("data_steps", []):
            if step.get("name") == name and not step.get("skip"):
                if step.get("x") is not None and step.get("y") is not None:
                    return step
        return None

    def teach(self):
        self._data = run_teach_mode()

    def _activate_dentweb(self, log_callback=None) -> bool:
        """'덴트웹' 창을 찾아서 포그라운드로 활성화"""
        def _log(msg):
            if log_callback:
                log_callback(msg)

        # 방법 1: pygetwindow
        try:
            import pygetwindow as gw
            all_windows = gw.getAllTitles()
            _log(f"열린 창 {len(all_windows)}개 탐색 중...")

            matched = [t for t in all_windows if WINDOW_TITLE in t]
            if not matched:
                _log(f"'{WINDOW_TITLE}' 포함된 창 없음")
                korean_wins = [t for t in all_windows if t.strip()]
                _log(f"전체 창 목록: {korean_wins[:10]}")
            else:
                _log(f"찾은 창: {matched[0]}")
                windows = gw.getWindowsWithTitle(matched[0])
                if windows:
                    win = windows[0]
                    if win.isMinimized:
                        win.restore()
                        time.sleep(0.5)
                    win.activate()
                    time.sleep(0.5)
                    if not win.isMaximized:
                        win.maximize()
                        time.sleep(0.5)
                    return True
        except Exception as e:
            _log(f"pygetwindow 오류: {e}")

        # 방법 2: win32gui (fallback)
        try:
            import ctypes
            import ctypes.wintypes

            def _find_window_callback(hwnd, results):
                if ctypes.windll.user32.IsWindowVisible(hwnd):
                    length = ctypes.windll.user32.GetWindowTextLengthW(hwnd)
                    if length > 0:
                        buf = ctypes.create_unicode_buffer(length + 1)
                        ctypes.windll.user32.GetWindowTextW(hwnd, buf, length + 1)
                        if WINDOW_TITLE in buf.value:
                            results.append(hwnd)
                return True

            WNDENUMPROC = ctypes.WINFUNCTYPE(
                ctypes.wintypes.BOOL, ctypes.wintypes.HWND, ctypes.wintypes.LPARAM
            )
            results = []
            ctypes.windll.user32.EnumWindows(
                WNDENUMPROC(lambda hwnd, _: _find_window_callback(hwnd, results) or True),
                0,
            )

            if results:
                hwnd = results[0]
                _log(f"win32gui로 창 발견 (hwnd={hwnd})")
                ctypes.windll.user32.ShowWindow(hwnd, 9)
                time.sleep(0.3)
                ctypes.windll.user32.SetForegroundWindow(hwnd)
                time.sleep(0.3)
                ctypes.windll.user32.ShowWindow(hwnd, 3)
                time.sleep(0.5)
                return True
            else:
                _log("win32gui로도 창을 찾을 수 없습니다")
        except Exception as e:
            _log(f"win32gui 오류: {e}")

        return False

    @staticmethod
    def excel_has_data(file_path: str) -> bool:
        """수술기록지 시트의 2행 이하에 데이터가 있는지 확인"""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet = None
            for name in wb.sheetnames:
                if "수술기록지" in name:
                    sheet = wb[name]
                    break
            if sheet is None:
                wb.close()
                return False
            has_data = False
            for row in sheet.iter_rows(min_row=2, max_row=2, max_col=1, values_only=True):
                if row[0] is not None and str(row[0]).strip():
                    has_data = True
                    break
            wb.close()
            return has_data
        except Exception:
            return False

    def _wait_for_download(self) -> str | None:
        """다운로드 폴더에서 엑셀 파일 감지"""
        target = os.path.join(self.download_dir, "dentweb_export.xlsx")
        deadline = time.time() + self.download_timeout
        before = set(glob.glob(os.path.join(self.download_dir, "*.xlsx")))

        while time.time() < deadline:
            if os.path.exists(target):
                time.sleep(1)
                return target
            current = set(glob.glob(os.path.join(self.download_dir, "*.xlsx")))
            new_files = current - before
            if new_files:
                time.sleep(1)
                return max(new_files, key=os.path.getmtime)
            time.sleep(1)
        return None

    def download_excel(self, log_callback=None) -> str | None:
        """전체 자동화: 덴트웹 활성화 → 엑셀 저장 → 데이터 유무 판별"""
        def _log(msg):
            if log_callback:
                log_callback(msg)

        if not self._data:
            _log("설정 데이터 없음")
            return None

        # 1. 덴트웹 창 자동 탐색 + 활성화
        _log("덴트웹 창 탐색 중...")
        if not self._activate_dentweb(log_callback=log_callback):
            _log("덴트웹 창을 찾을 수 없습니다")
            return None
        _log("덴트웹 창 활성화 완료 — 2초 대기")
        time.sleep(2)

        # 2. 날짜 선택 시퀀스 (엑셀저장 전까지)
        is_first_click = True
        for step in self._data.get("data_steps", []):
            if step.get("skip"):
                continue
            if step.get("name") == "data_check":
                continue
            if step.get("name") == "export_btn":
                break
            if not _find_and_click(step, log_callback, activate_first=is_first_click):
                _log(f"단계 실패: {step['label']}")
                return None
            is_first_click = False  # 이후 단계는 이미 활성화됨
            time.sleep(step.get("wait_after", 1.5))

        # 3. 엑셀저장 클릭
        export_step = None
        for step in self._data.get("data_steps", []):
            if step.get("name") == "export_btn" and not step.get("skip"):
                export_step = step
                break
        if not export_step:
            _log("엑셀저장 버튼 설정 없음")
            return None

        if not _find_and_click(export_step, log_callback):
            _log("엑셀저장 버튼 클릭 실패")
            return None
        time.sleep(export_step.get("wait_after", 3.0))

        # 4. "다른 이름으로 저장" 다이얼로그 처리
        _log("저장 다이얼로그 처리 중...")

        # 4a. '덴트웹 에이전트' 폴더 더블클릭
        agent_step = self._get_save_step("save_dialog_agent_folder")
        if agent_step:
            _log("'덴트웹 에이전트' 폴더 더블클릭")
            pyautogui.doubleClick(int(agent_step["x"]), int(agent_step["y"]))
            time.sleep(agent_step.get("wait_after", 1.5))
        else:
            _log("[경고] 덴트웹 에이전트 폴더 좌표 미설정 — 폴더 탐색 생략")

        # 4b. 'exports' 폴더 더블클릭
        exports_step = self._get_save_step("save_dialog_exports_folder")
        if exports_step:
            _log("'exports' 폴더 더블클릭")
            pyautogui.doubleClick(int(exports_step["x"]), int(exports_step["y"]))
            time.sleep(exports_step.get("wait_after", 1.5))
        else:
            _log("[경고] exports 폴더 좌표 미설정 — 폴더 탐색 생략")

        # 4c. 파일명 필드 클릭 → 전체 선택 → 파일명 입력
        filename_step = self._get_save_step("save_dialog_filename_field")
        if filename_step:
            _log("파일 이름 필드 클릭")
            _win32_click(int(filename_step["x"]), int(filename_step["y"]))
            time.sleep(filename_step.get("wait_after", 0.5))
        else:
            _log("[경고] 파일 이름 필드 좌표 미설정 — Alt+N으로 폴백")
            pyautogui.hotkey("alt", "n")
            time.sleep(0.3)
        pyautogui.hotkey("ctrl", "a")
        time.sleep(0.2)
        _paste_text("dentweb_export")
        time.sleep(0.3)

        # 4d. '저장(S)' 버튼 클릭
        save_btn_step = self._get_save_step("save_dialog_save_btn")
        if save_btn_step:
            _log("'저장(S)' 버튼 클릭")
            _win32_click(int(save_btn_step["x"]), int(save_btn_step["y"]))
            time.sleep(save_btn_step.get("wait_after", 2.0))
        else:
            _log("[경고] 저장 버튼 좌표 미설정 — Enter로 폴백")
            pyautogui.press("enter")
            time.sleep(2)

        # 5. 덮어쓰기 확인 팝업 — '예(Y)' 클릭
        confirm_step = self._get_save_step("save_dialog_confirm_yes")
        if confirm_step:
            _log("덮어쓰기 확인 '예(Y)' 클릭")
            _win32_click(int(confirm_step["x"]), int(confirm_step["y"]))
            time.sleep(confirm_step.get("wait_after", 2.0))
        else:
            _log("[경고] 예(Y) 버튼 좌표 미설정 — Enter로 폴백")
            pyautogui.press("enter")
            time.sleep(2)

        # 6. 파일 저장 대기
        _log("엑셀 파일 저장 대기 중...")
        excel_path = self._wait_for_download()
        if not excel_path:
            _log("엑셀 파일 저장 시간 초과")
            return None
        _log(f"엑셀 파일 저장 완료: {os.path.basename(excel_path)}")

        # 7. "수술기록지" 시트 2행 데이터 확인
        _log("수술기록지 데이터 확인 중...")
        if not self.excel_has_data(excel_path):
            _log("수술기록지에 데이터 없음 (빈 파일)")
            self.cleanup(excel_path)
            return None

        _log("수술 기록 데이터 확인 완료")
        return excel_path

    def cleanup(self, file_path: str):
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except OSError:
            pass
